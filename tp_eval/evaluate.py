#!/usr/bin/env python2

import argparse
import datetime
from itertools import product, combinations
import logging as log
import math
import matplotlib.pyplot as plt
from mpl_toolkits import mplot3d
import numpy as np
import openpyxl
import os
import px4tools
import re
import sys
import xlsxwriter



########


# get_ipython().magic(u'matplotlib inline')


def mean(numbers):
    return float(sum(numbers)) / max(len(numbers), 1)


def find_string(pat, text):

    match = re.search(pat, text)
    if match:
        return match.group()
    else:
        return('not found')


def get_data_from_log(settings):
    # Folder of this script
    source_folder = os.path.dirname(os.path.abspath(__file__))

    if settings['dev']:
        # os.chdir('../build_posix_sitl_dronecourse_dev/tmp/rootfs/fs/microsd/log')
        log_folder = source_folder + '/../build_posix_sitl_dronecourse_dev/tmp/rootfs/fs/microsd/log'
    else:
        # os.chdir('../build_posix_sitl_dronecourse/tmp/rootfs/fs/microsd/log')
        log_folder = source_folder + '/../build_posix_sitl_dronecourse/tmp/rootfs/fs/microsd/log'

    # List all logged sessions
    session_prop = []
    session_folders = os.listdir(log_folder)    
    for session in session_folders:
        # append only if it is called 'sess#'
        if 'sess' in session:
            path_to_session = log_folder + '/' + session 
            session_prop.append((path_to_session, 
                                 os.stat(path_to_session).st_mtime))
    log.info("sessions found : ", session_prop)

    # Find newest session folder
    session_prop.sort(key=lambda tup: tup[1])
    newest_session_folder = session_prop[-1][0]

    # Find log files in this folder
    logfiles = [x for x in os.listdir(newest_session_folder) if '.ulg' in x]
    log.info("files found in logger : ", logfiles)

    # Read log    
    d = px4tools.read_ulog(newest_session_folder + '/' + logfiles[0])
    data = d.concat(dt=0.01)

    # Clean dataset
    data = data[np.isfinite(data['t_vehicle_local_position_groundtruth_0__f_x'])]
    data = data[np.isfinite(data['t_vehicle_local_position_groundtruth_0__f_y'])]
    data = data[np.isfinite(data['t_vehicle_local_position_groundtruth_0__f_z'])]
    data = data[np.isfinite(data['timestamp'])]

    print('Data was analyzed from file ' + newest_session_folder + '/' + logfiles[0] + '\n\n')

    return data


def evaluation_task_1(data, perf, settings, info):

    WAYPOINT_ACTIVATION_RANGE = 6
    WAYPOINT_FULLSCORE_RANGE = 2
    WAYPOINT_MAX_SCORE = 100

    EVAL_M = -WAYPOINT_MAX_SCORE/(WAYPOINT_ACTIVATION_RANGE - WAYPOINT_FULLSCORE_RANGE)
    EVAL_Q = WAYPOINT_MAX_SCORE - WAYPOINT_FULLSCORE_RANGE * EVAL_M

    # check timeout

    if 't_dronecourse_timeouts_0__f_timeout_task1' in list(data):
        timeout = data[np.isfinite(data['t_dronecourse_timeouts_0__f_timeout_task1'])]
        timeout = timeout["t_dronecourse_timeouts_0__f_timeout_task1"][0]
    else:
        timeout = False


    data_task1 = data[data['t_dronecourse_status_0__f_mode'] == 1]

    if not len(data_task1):
        print 'no data found for task 1'
        perf['task1_not_done'] = True
        return

    px = data_task1["t_vehicle_local_position_groundtruth_0__f_x"]
    py = data_task1["t_vehicle_local_position_groundtruth_0__f_y"]
    pz = data_task1["t_vehicle_local_position_groundtruth_0__f_z"]
    ts = data_task1["timestamp"]

    # #

    pxarr = -px.values
    pyarr = py.values
    pzarr = -pz.values
    tarr = ts.values/1e6  # timestamp is in us


    perf['x_task1'] = pxarr
    perf['y_task1'] = pyarr
    perf['z_task1'] = pzarr
    perf['t_task1'] = tarr

    # Get waypoints

    data_task1_wp = data[np.isfinite(data['t_dronecourse_waypoint_0__f_x1'])]

    wp1_x = data_task1_wp["t_dronecourse_waypoint_0__f_x1"]
    wp1_y = data_task1_wp["t_dronecourse_waypoint_0__f_y1"]
    wp1_z = data_task1_wp["t_dronecourse_waypoint_0__f_z1"]

    wp2_x = data_task1_wp["t_dronecourse_waypoint_0__f_x2"]
    wp2_y = data_task1_wp["t_dronecourse_waypoint_0__f_y2"]
    wp2_z = data_task1_wp["t_dronecourse_waypoint_0__f_z2"]

    wp3_x = data_task1_wp["t_dronecourse_waypoint_0__f_x3"]
    wp3_y = data_task1_wp["t_dronecourse_waypoint_0__f_y3"]
    wp3_z = data_task1_wp["t_dronecourse_waypoint_0__f_z3"]

    pwx_t = [wp1_x.values[-1], wp2_x.values[-1], wp3_x.values[-1]]
    pwy = [wp1_y.values[-1], wp2_y.values[-1], wp3_y.values[-1]]
    pwz_t = [wp1_z.values[-1], wp2_z.values[-1], wp3_z.values[-1]]
    pwx = [x*-1 for x in pwx_t]
    pwz = [x*-1 for x in pwz_t]

    perf['pwx'] = pwx
    perf['pwy'] = pwy
    perf['pwz'] = pwz

    # print pwx

    # print settings

    if settings['write_txt'] or settings['write_spreadsheet']:
        file_n = 'TP_1_eval_of_{}'.format(info['date_time'])

    if settings['write_spreadsheet']:

        # Fill spreadsheet

        # Create a workbook and add a worksheet.

        workbook = xlsxwriter.Workbook(file_n + '.xlsx')
        worksheet = workbook.add_worksheet()

        # Add a bold format to use to highlight cells.

        bold = workbook.add_format({'bold': True})

        # Write some data headers.

        worksheet.write('A1', 'time [s]', bold)
        worksheet.write('B1', 'x coordinate [m]', bold)
        worksheet.write('C1', 'y coordinate [m]', bold)
        worksheet.write('D1', 'z coordinate [m]', bold)

        worksheet.write('G1', 'x (waypoints) [m]', bold)
        worksheet.write('H1', 'y (waypoints) [m]', bold)
        worksheet.write('I1', 'z (waypoints) [m]', bold)


        # Start from the first cell below the headers.

        row = 1
        col = 6

        # Iterate over the data and write it out row by row.
        for item in range(len(pwx)):
            worksheet.write(row, 5,     item+1, bold)
            worksheet.write(row, col,     pwx[item])
            worksheet.write(row, col + 1,     pwy[item])
            worksheet.write(row, col + 2,     pwz[item])
            row += 1

        # Start from the first cell below the headers.
        row = 1
        col = 0


        # Iterate over the data and write it out row by row.
        for item in range(len(pxarr.tolist())):
            worksheet.write(row, col,     tarr.tolist()[item] if not math.isnan(tarr.tolist()[item]) else 0)
            worksheet.write(row, col + 1,     pxarr.tolist()[item] if not math.isnan(pxarr.tolist()[item]) else 0)
            worksheet.write(row, col + 2,     pyarr.tolist()[item] if not math.isnan(pyarr.tolist()[item]) else 0)
            worksheet.write(row, col + 3,     pzarr.tolist()[item] if not math.isnan(pzarr.tolist()[item]) else 0)
            row += 1

        workbook.close()

    ##

    minval = []
    minidx = []

    if settings['write_txt']:
        file_name = open(file_n + '.txt', 'w')
        # print str0
        # file_name.write(str0)

    score = []

    fig = plt.figure("Waypoint Navigation")

    for i in xrange(3):
        xdist = pxarr-pwx[i]
        ydist = pyarr-pwy[i]
        zdist = pzarr-pwz[i]

        dist = np.sqrt(xdist**2+ydist**2+zdist**2)

        minval.append(np.amin(dist))
        minidx.append(np.argmin(dist))

        dist_cross = dist[dist < WAYPOINT_ACTIVATION_RANGE]
        t_cross = tarr[dist < WAYPOINT_ACTIVATION_RANGE]

        if minval[i] < WAYPOINT_FULLSCORE_RANGE:
            score.append(100)
        else:
            score.append(minval[i] * EVAL_M + EVAL_Q)

        if len(t_cross) > 0:
            str1 = 'you crossed waypoint n. {0} after {1:.2f} seconds'.format(i+1, t_cross[0])
            str2 = 'minimum distance from waypoint n. {0} is {1:.2f} meters after {2:.2f} seconds'.format(i+1, minval[i], tarr[minidx[i]])
            str3 = 'you scored {0:.2f} accuracy points'.format(score[i])

            perf['cross_time_waypoints'].append(t_cross[0])
            perf['min_dist_waypoints'].append(minval[i])
            perf['score_waypoints'].append(score[i])

        else:
            score[i] = 0
            perf['score_waypoints'].append(score[i])
            perf['cross_time_waypoints'].append(600)
            str1 = 'you did not cross waypoint n. {0} before timeout'.format(i+1)
            str2 = 'you scored 0 accuracy points'
            str3 = ''

        print str1
        print str2
        print str3
        print '\n'

        if settings['write_txt']:
            file_name.write(str1 + '\n')
            file_name.write(str2 + '\n')
            file_name.write(str3 + '\n')
            file_name.write('\n')

        if settings['show_figures']:
            ax = fig.add_subplot(1, 3, i+1)
            ax.set_xlabel('time [s]')
            ax.set_ylabel('distance from waypoint n.{} [m]'.format(i+1))
            ax.grid()
            ax.plot(tarr, dist)


    if len(score) >= 1:
        perf['accuracy_score_task1'] = mean(perf['score_waypoints'])
        perf['time_task1'] = perf['cross_time_waypoints'][-1] + 10

    if perf['time_task1'] > 600:
        perf['time_task1'] = 600

    if settings['write_txt']:
        file_name.close()

    #

    x = np.linspace(-10, 10, 100)
    y = np.linspace(-10, 10, 100)
    z = np.linspace(-10, 10, 100)

    #

    if settings['show_figures']:
        fig = plt.figure("Waypoint Navigation - Trajectory")
    
        ax = fig.add_subplot(1, 1, 1, projection='3d')
    
        ax.plot(pxarr, pyarr, zs=pzarr)


    x_m = mean([np.amin(pxarr), np.amax(pxarr)])
    y_m = mean([np.amin(pyarr), np.amax(pyarr)])
    z_m = mean([np.amin(pzarr), np.amax(pzarr)])

    x_range = max(np.abs(np.amin(pxarr) - x_m), np.abs(np.amax(pxarr) - x_m))
    y_range = max(np.abs(np.amin(pyarr) - y_m), np.abs(np.amax(pyarr) - y_m))
    z_range = max(np.abs(np.amin(pzarr) - z_m), np.abs(np.amax(pzarr) - z_m))

    range_fin = max([x_range, y_range, z_range])

    x_range = [x_m-range_fin/2, x_m+range_fin/2]
    y_range = [y_m-range_fin/2, y_m+range_fin/2]
    z_range = [z_m-range_fin/2, z_m+range_fin/2]

    # plot ground level

    # X = np.arange(-range_fin, range_fin, range_fin*2-1)
    # Y = np.arange(-range_fin, range_fin, range_fin*2-1)
    # X, Y = np.meshgrid(X, Y)
    # Z = 0*(X+Y)
    # surf = ax.plot_surface(X, Y, Z, color = 'y')

    # # plot spheres

    if settings['show_figures']:
        for i in xrange(3):
            u, v = np.mgrid[0:2*np.pi:20j, 0:np.pi:10j]
            x = pwx[i] + np.cos(u)*np.sin(v)
            y = pwy[i] + np.sin(u)*np.sin(v)
            z = pwz[i] + np.cos(v)
            ax.plot_wireframe(x, y, z, color="r")

        ax.set_xlim(x_range)
        ax.set_ylim(y_range)
        ax.set_zlim(z_range)
    
        ax.set_xlabel('x')
        ax.set_ylabel('y')
        ax.set_zlabel('z')


def evaluation_task_2(data, perf, settings, info):

    PLATFORM_MAX_SCORE = 100

    EDGE_SIZE = 1
    FULL_SCORE_EDGE_SIZE = 0.2

    # LIMIT_DISTANCE_FOR_ANALYSIS = 10

    EVAL_M = -PLATFORM_MAX_SCORE/(EDGE_SIZE - FULL_SCORE_EDGE_SIZE)
    EVAL_Q = PLATFORM_MAX_SCORE - FULL_SCORE_EDGE_SIZE * EVAL_M

    # check timeout

    if 't_dronecourse_timeouts_0__f_timeout_task2' in list(data):
        timeout = data[np.isfinite(data['t_dronecourse_timeouts_0__f_timeout_task2'])]
        timeout = timeout["t_dronecourse_timeouts_0__f_timeout_task2"][0]
    else:
        timeout = False

    if timeout:
        print 'timeout during task 2'
        print ''
        perf['accuracy_score_task2'] = 0
        perf['time_task2'] = 600
        return

    data_task2_pl = data[np.isfinite(data['t_dronecourse_platform_position_0__f_x'])]

    data_task2 = data[data['t_dronecourse_status_0__f_mode'] == 2]

    if not len(data_task2):
        print 'no data found for task 2 \n'
        perf['task2_not_done'] = True
        return

    # Clean dataset

    data_task2 = data_task2[np.isfinite(data_task2['t_vehicle_local_position_groundtruth_0__f_x'])]
    data_task2 = data_task2[np.isfinite(data_task2['t_vehicle_local_position_groundtruth_0__f_y'])]
    data_task2 = data_task2[np.isfinite(data_task2['t_vehicle_local_position_groundtruth_0__f_z'])]
    data_task2 = data_task2[np.isfinite(data_task2['timestamp'])]

    # data_task2 = data_task2[abs(data_task2['t_vehicle_local_position_groundtruth_0__f_x']-plat_x) < LIMIT_DISTANCE_FOR_ANALYSIS]
    # data_task2 = data_task2[abs(data_task2['t_vehicle_local_position_groundtruth_0__f_y']-plat_y) < LIMIT_DISTANCE_FOR_ANALYSIS]

    # #

    px2 = data_task2["t_vehicle_local_position_groundtruth_0__f_x"]
    py2 = data_task2["t_vehicle_local_position_groundtruth_0__f_y"]
    pz2 = data_task2["t_vehicle_local_position_groundtruth_0__f_z"]
    landed = data_task2['t_vehicle_land_detected_0__f_landed']
    ts2 = data_task2["timestamp"]

     

    pl_x = data_task2_pl["t_dronecourse_platform_position_0__f_x"]
    pl_y = data_task2_pl["t_dronecourse_platform_position_0__f_y"]
    pl_z = data_task2_pl["t_dronecourse_platform_position_0__f_z"]
    
    plat_x = -pl_x.values[-1]
    plat_y = pl_y.values[-1]
    plat_z = -pl_z.values[-1]
    
    perf['platform'] = [plat_x, plat_y, plat_z]

    # #

    pxarr2 = -px2.values
    pyarr2 = py2.values
    pzarr2 = -pz2.values
    landedarr = landed.values
    tarr2 = ts2.values/1e6  # timestamp is in us

    if not sum(landedarr == 1):
        print 'no landing detected during task 2'
        print ''
        perf['task2_not_done'] = True
        return


    perf['x_task2'] = pxarr2
    perf['y_task2'] = pyarr2
    perf['z_task2'] = pzarr2
    perf['t_task2'] = tarr2

    # #

    if settings['show_figures']:
        fig = plt.figure("Platform landing")
    
        ax = fig.add_subplot(1, 2, 1, projection='3d')
    
        ax.plot(pxarr2, pyarr2, zs=pzarr2)

    x_m2 = mean([np.amin(pxarr2), np.amax(pxarr2)])
    y_m2 = mean([np.amin(pyarr2), np.amax(pyarr2)])
    z_m2 = mean([np.amin(pzarr2), np.amax(pzarr2)])

    x_range2 = max(np.abs(np.amin(pxarr2) - x_m2), np.abs(np.amax(pxarr2) - x_m2))
    y_range2 = max(np.abs(np.amin(pyarr2) - y_m2), np.abs(np.amax(pyarr2) - y_m2))
    z_range2 = max(np.abs(np.amin(pzarr2) - z_m2), np.abs(np.amax(pzarr2) - z_m2))

    range_fin2 = max([x_range2, y_range2, z_range2])

    x_range2 = [x_m2 - range_fin2 / 2, x_m2 + range_fin2 / 2]
    y_range2 = [y_m2 - range_fin2 / 2, y_m2 + range_fin2 / 2]
    z_range2 = [z_m2 - range_fin2 / 2, z_m2 + range_fin2 / 2]

    # # plot 3D platform

    r1 = [-EDGE_SIZE, EDGE_SIZE]
    r2 = [-EDGE_SIZE, EDGE_SIZE]
    r3 = [-plat_z/2, plat_z/2]
    center = [plat_x, plat_y, plat_z/2]

    for s, e in combinations(np.array(list(product(r1, r2, r3))), 2):
        s = np.array(center)+np.array(s)
        e = np.array(center)+np.array(e)
        if np.linalg.norm(s-e) == 2*r1[1] or np.linalg.norm(s-e) == 2*r2[1] or np.linalg.norm(s-e) == 2*r3[1]:
            if settings['show_figures']:
                ax.plot3D(*zip(s, e), color="r")

    if settings['show_figures']:
        ax.set_xlim(x_range2)
        ax.set_ylim(y_range2)
        ax.set_zlim(z_range2)
    
        ax.set_xlabel('x')
        ax.set_ylabel('y')
        ax.set_zlabel('z')

    # # plot 2D platform + landing spot

    if settings['show_figures']:
        ax = fig.add_subplot(1, 2, 2)

    # # plot 2D platform

    if settings['show_figures']:
        ax.plot([- EDGE_SIZE, + EDGE_SIZE], [- EDGE_SIZE, - EDGE_SIZE], 'r', label='platform edge')
        ax.plot([- EDGE_SIZE, + EDGE_SIZE], [+ EDGE_SIZE, + EDGE_SIZE], 'r')
        ax.plot([+ EDGE_SIZE, + EDGE_SIZE], [- EDGE_SIZE, + EDGE_SIZE], 'r')
        ax.plot([- EDGE_SIZE, - EDGE_SIZE], [- EDGE_SIZE, + EDGE_SIZE], 'r')
    
        ax.plot([- FULL_SCORE_EDGE_SIZE, + FULL_SCORE_EDGE_SIZE], [- FULL_SCORE_EDGE_SIZE, - FULL_SCORE_EDGE_SIZE], 'C2', label='full score range')
        ax.plot([- FULL_SCORE_EDGE_SIZE, + FULL_SCORE_EDGE_SIZE], [+ FULL_SCORE_EDGE_SIZE, + FULL_SCORE_EDGE_SIZE], 'C2')
        ax.plot([+ FULL_SCORE_EDGE_SIZE, + FULL_SCORE_EDGE_SIZE], [- FULL_SCORE_EDGE_SIZE, + FULL_SCORE_EDGE_SIZE], 'C2')
        ax.plot([- FULL_SCORE_EDGE_SIZE, - FULL_SCORE_EDGE_SIZE], [- FULL_SCORE_EDGE_SIZE, + FULL_SCORE_EDGE_SIZE], 'C2')

    # landing spot

    t_land2 = tarr2[landedarr == 1][0]

    land_x2 = pxarr2[landedarr == 1][0]
    land_y2 = pyarr2[landedarr == 1][0]

    if settings['show_figures']:
        ax.scatter(land_x2 - plat_x, land_y2 - plat_y, s=100, c='b', label='landing position')

    # print land_x2 - plat_x
    # print land_y2 - plat_y

    dist2 = (max([abs(land_x2 - plat_x), abs(land_y2 - plat_y)]))
    
    if settings['show_figures']:
        ax.plot([- dist2, + dist2], [- dist2, - dist2], 'b--')
        ax.plot([- dist2, + dist2], [+ dist2, + dist2], 'b--')
        ax.plot([+ dist2, + dist2], [- dist2, + dist2], 'b--')
        ax.plot([- dist2, - dist2], [- dist2, + dist2], 'b--')
    
        ax.legend(bbox_to_anchor=(0., 1.02, 1., .102), loc=3,
               ncol=1, mode="expand", borderaxespad=0.)
        ax.grid()

    if dist2 < FULL_SCORE_EDGE_SIZE:
        score_plat = 100
    elif dist2 > EDGE_SIZE:
        score_plat = 0
    else:
        score_plat = (dist2*EVAL_M+EVAL_Q)

    if 1:  # managed
        str1 = 'you landed on the platform after after {0:.2f} seconds'.format(t_land2)  # t_cross[0])
        str2 = 'L_inf distance from platform center is {0:.2f} meters'.format(dist2)
        str3 = 'you scored {0:.2f} accuracy points'.format(score_plat)

        perf['accuracy_score_task2'] = score_plat
        perf['time_task2'] = t_land2 - perf['time_task1']
    else:
        str1 = 'you did not manage to land on the plaform before timeout'
        str2 = 'you scored 0 accuracy points'
        str3 = ''

    print '\n'
    print '\n'
    print str1
    print str2
    print str3
    print '\n'

    if settings['write_txt'] or settings['write_spreadsheet']:
        file_n = 'TP_2_eval_of_{}'.format(info['date_time'])

    if settings['write_txt']:

        file_name = open(file_n + '.txt', 'w')
        # file_name.write(str0)

        file_name.write('\n')
        file_name.write('\n')
        file_name.write(str1 + '\n')
        file_name.write(str2 + '\n')
        file_name.write(str3 + '\n')
        file_name.write('\n')

        file_name.close()

    if settings['write_spreadsheet']:
        # Create a workbook and add a worksheet.

        workbook = xlsxwriter.Workbook(file_n + '.xlsx')
        worksheet = workbook.add_worksheet()

        # Add a bold format to use to highlight cells.

        bold = workbook.add_format({'bold': True})

        # Write some data headers.

        worksheet.write('A1', 'time [s]', bold)
        worksheet.write('B1', 'x coordinate [m]', bold)
        worksheet.write('C1', 'y coordinate [m]', bold)
        worksheet.write('D1', 'z coordinate [m]', bold)

        worksheet.write('G1', 'x (plaform center) [m]', bold)
        worksheet.write('H1', 'y (plaform center) [m]', bold)
        worksheet.write('I1', 'z (plaform center) [m]', bold)


        # Start from the first cell below the headers.

        row = 1
        col = 6

        # Iterate over the data and write it out row by row.
        worksheet.write(row, col,     plat_x)
        worksheet.write(row, col + 1,     plat_y)
        worksheet.write(row, col + 2,     plat_z)

        # Start from the first cell below the headers.
        row = 1
        col = 0


        # Iterate over the data and write it out row by row.
        for item in range(len(pxarr2.tolist())):
            worksheet.write(row, col,     tarr2.tolist()[item] if not math.isnan(tarr2.tolist()[item]) else 0)
            worksheet.write(row, col + 1,     pxarr2.tolist()[item] if not math.isnan(pxarr2.tolist()[item]) else 0)
            worksheet.write(row, col + 2,     pyarr2.tolist()[item] if not math.isnan(pyarr2.tolist()[item]) else 0)
            worksheet.write(row, col + 3,     pzarr2.tolist()[item] if not math.isnan(pzarr2.tolist()[item]) else 0)
            row += 1

        workbook.close()


def evaluation_task_3(data, perf, settings, info):

    PLATFORM_MAX_SCORE = 100

    EDGE_SIZE_Y = 2
    EDGE_SIZE_X = 1.05
    FULL_SCORE_EDGE_SIZE_Y = 0.4
    FULL_SCORE_EDGE_SIZE_X = 0.2

    LIMIT_DISTANCE_FOR_ANALYSIS = 10


    # check timeout

    if 't_dronecourse_timeouts_0__f_timeout_task2' in list(data):
        timeout = data[np.isfinite(data['t_dronecourse_timeouts_0__f_timeout_task3'])]
        timeout = timeout["t_dronecourse_timeouts_0__f_timeout_task3"][0]
    else:
        timeout = False

    if timeout:
        print 'timeout during task 3'
        print ''
        perf['accuracy_score_task3'] = 0
        perf['time_task3'] = 600
        return


    data_task3 = data[data['t_dronecourse_status_0__f_mode'] == 3]

    if not len(data_task3):
        print 'no data found for task 3'
        print ''
        perf['task3_not_done'] = True
        return

    # Clean dataset

    data_task3 = data_task3[np.isfinite(data_task3['t_vehicle_local_position_groundtruth_0__f_x'])]
    data_task3 = data_task3[np.isfinite(data_task3['t_vehicle_local_position_groundtruth_0__f_y'])]
    data_task3 = data_task3[np.isfinite(data_task3['t_vehicle_local_position_groundtruth_0__f_z'])]
    data_task3 = data_task3[np.isfinite(data_task3['timestamp'])]
    
    data_task3 = data_task3[np.isfinite(data_task3["t_dronecourse_truck_position_0__f_x"])]
    data_task3 = data_task3[np.isfinite(data_task3["t_dronecourse_truck_position_0__f_y"])]
    data_task3 = data_task3[np.isfinite(data_task3["t_dronecourse_truck_position_0__f_z"])]
    
    data_task3 = data_task3[100:]   # very dirty solution for first takeoff!

    # #

    px = data_task3["t_vehicle_local_position_groundtruth_0__f_x"]
    py = data_task3["t_vehicle_local_position_groundtruth_0__f_y"]
    pz = data_task3["t_vehicle_local_position_groundtruth_0__f_z"]

    landed3 = data_task3['t_vehicle_land_detected_0__f_landed']

    ts = data_task3["timestamp"]

    tr_x = data_task3["t_dronecourse_truck_position_0__f_x"]
    tr_y = data_task3["t_dronecourse_truck_position_0__f_y"]
    tr_z = data_task3["t_dronecourse_truck_position_0__f_z"]

    tr_vx = data_task3["t_dronecourse_truck_position_0__f_vx"]
    tr_vy = data_task3["t_dronecourse_truck_position_0__f_vy"]
    tr_vz = data_task3["t_dronecourse_truck_position_0__f_vz"]


    if 't_target_position_ned_filtered_0__f_x' not in list(data_task3):
        print 'kalman filter not implemented \n'
        print ''
        perf['task3_not_done'] = True
        return

    tr_x_eval = data_task3["t_target_position_ned_filtered_0__f_x"]
    tr_y_eval = data_task3["t_target_position_ned_filtered_0__f_y"]
    tr_z_eval = data_task3["t_target_position_ned_filtered_0__f_z"]
    
    tr_vx_eval = data_task3["t_target_position_ned_filtered_0__f_vx"]
    tr_vy_eval = data_task3["t_target_position_ned_filtered_0__f_vy"]
    tr_vz_eval = data_task3["t_target_position_ned_filtered_0__f_vz"]

    tr_x_meas = data_task3["t_target_position_ned_0__f_x"]
    tr_y_meas = data_task3["t_target_position_ned_0__f_y"]
    tr_z_meas = data_task3["t_target_position_ned_0__f_z"]

    # #

    pxarr = -px.values
    pyarr = py.values
    pzarr = -pz.values

    perf['x_task3'] = pxarr
    perf['y_task3'] = pyarr
    perf['z_task3'] = pzarr

    #

    # tr_xarr = -tr_x.values
    tr_xarr = -tr_x.values
    tr_yarr = tr_y.values
    tr_zarr = -tr_z.values

    tr_xarr_eval = -tr_x_eval.values
    tr_yarr_eval = tr_y_eval.values
    tr_zarr_eval = -tr_z_eval.values

    tr_xarr_meas = -tr_x_meas.values
    tr_yarr_meas = tr_y_meas.values
    tr_zarr_meas = -tr_z_meas.values


    perf['x_tr'] = tr_xarr
    perf['y_tr'] = tr_yarr
    perf['z_tr'] = tr_zarr

    perf['x_tr_eval'] = tr_xarr_eval
    perf['y_tr_eval'] = tr_yarr_eval
    perf['z_tr_eval'] = tr_zarr_eval

    perf['x_tr_meas'] = tr_xarr_meas
    perf['y_tr_meas'] = tr_yarr_meas
    perf['z_tr_meas'] = tr_zarr_meas

    #

    tr_vxarr = -tr_vx.values
    tr_vyarr = tr_vy.values
    tr_vzarr = -tr_vz.values

    tr_vxarr_eval = -tr_vx_eval.values
    tr_vyarr_eval = tr_vy_eval.values
    tr_vzarr_eval = -tr_vz_eval.values


    perf['xv_tr'] = tr_vxarr
    perf['yv_tr'] = tr_vyarr
    perf['zv_tr'] = tr_vzarr

    perf['vx_tr_eval'] = tr_vxarr_eval
    perf['vy_tr_eval'] = tr_vyarr_eval
    perf['vz_tr_eval'] = tr_vzarr_eval

    landedarr = landed3.values
    tarr = ts.values/1e6  # timestamp is in us

    perf['t_task3'] = tarr

    if not sum(landedarr == 1):
        print 'no landing detected during task 3'
        print ''
        perf['task3_not_done'] = True
        return

    # #

    if settings['show_figures']:
        fig = plt.figure("Truck landing")
    
        ax = fig.add_subplot(1, 2, 1, projection='3d')
    
        ax.plot(pxarr, pyarr, zs=pzarr)

    x_m3 = mean([np.amin(pxarr), np.amax(pxarr)])
    y_m3 = mean([np.amin(pyarr), np.amax(pyarr)])
    z_m3 = mean([np.amin(pzarr), np.amax(pzarr)])

    x_range3 = max(np.abs(np.amin(pxarr) - x_m3), np.abs(np.amax(pxarr) - x_m3))
    y_range3 = max(np.abs(np.amin(pyarr) - y_m3), np.abs(np.amax(pyarr) - y_m3))
    z_range3 = max(np.abs(np.amin(pzarr) - z_m3), np.abs(np.amax(pzarr) - z_m3))

    range_fin3 = max([x_range3, y_range3, z_range3])

    x_range3 = [x_m3 - range_fin3 / 2, x_m3 + range_fin3 / 2]
    y_range3 = [y_m3 - range_fin3 / 2, y_m3 + range_fin3 / 2]
    z_range3 = [z_m3 - range_fin3 / 2, z_m3 + range_fin3 / 2]

    # # plot truck position
    
    if settings['show_figures']:
        ax.plot(tr_xarr, tr_yarr, zs=tr_zarr)
    
        ax.set_xlim(x_range3)
        ax.set_ylim(y_range3)
        ax.set_zlim(z_range3)
    
        ax.set_xlabel('x')
        ax.set_ylabel('y')
        ax.set_zlabel('z')

        # # plot 2D platform + landing spot
    
        ax = fig.add_subplot(1, 2, 2)
    
        # # plot 2D truck surface
    
        ax.plot([- EDGE_SIZE_X, + EDGE_SIZE_X], [- EDGE_SIZE_Y, - EDGE_SIZE_Y], 'r', label='platform edge')
        ax.plot([- EDGE_SIZE_X, + EDGE_SIZE_X], [+ EDGE_SIZE_Y, + EDGE_SIZE_Y], 'r')
        ax.plot([+ EDGE_SIZE_X, + EDGE_SIZE_X], [- EDGE_SIZE_Y, + EDGE_SIZE_Y], 'r')
        ax.plot([- EDGE_SIZE_X, - EDGE_SIZE_X], [- EDGE_SIZE_Y, + EDGE_SIZE_Y], 'r')
    
        ax.plot([- FULL_SCORE_EDGE_SIZE_X, + FULL_SCORE_EDGE_SIZE_X], [- FULL_SCORE_EDGE_SIZE_Y, - FULL_SCORE_EDGE_SIZE_Y], 'C2', label='full score range')
        ax.plot([- FULL_SCORE_EDGE_SIZE_X, + FULL_SCORE_EDGE_SIZE_X], [+ FULL_SCORE_EDGE_SIZE_Y, + FULL_SCORE_EDGE_SIZE_Y], 'C2')
        ax.plot([+ FULL_SCORE_EDGE_SIZE_X, + FULL_SCORE_EDGE_SIZE_X], [- FULL_SCORE_EDGE_SIZE_Y, + FULL_SCORE_EDGE_SIZE_Y], 'C2')
        ax.plot([- FULL_SCORE_EDGE_SIZE_X, - FULL_SCORE_EDGE_SIZE_X], [- FULL_SCORE_EDGE_SIZE_Y, + FULL_SCORE_EDGE_SIZE_Y], 'C2')

    # landing spot
        
    islanded3 = landedarr == 1
        
    t_land3 = tarr[islanded3][0]

    land_x = pxarr[islanded3][0]
    land_y = pyarr[islanded3][0]
    
    tr_x = tr_xarr[islanded3][0]
    tr_y = tr_yarr[islanded3][0]
    
    tr_vx = tr_vxarr[islanded3][0]
    tr_vy = tr_vyarr[islanded3][0]
    
    tr_vnorm = np.sqrt(tr_vx*tr_vx + tr_vy*tr_vy)
    
    # print 'vx truck =', tr_vx
    # print 'vy truck =', tr_vy
    # print 'v norm truck =', tr_vnorm
    
    tr_tilt_angle = np.arccos(tr_vy/tr_vnorm)

    # print tr_vx/tr_vnorm
    # print tr_tilt_angle
    # print tr_tilt_angle/3.141592*180
    
    dist_x = land_x - tr_x
    dist_y = land_y - tr_y
    
    dist3 = abs(max([dist_x, dist_y]))
   
    dist_x_rot = dist_x * np.cos(tr_tilt_angle) - dist_y * np.sin(tr_tilt_angle)
    dist_y_rot = dist_x * np.sin(tr_tilt_angle) + dist_y * np.cos(tr_tilt_angle)
    
    # print dist_x
    # print dist_y
    # print dist_x_rot
    # print dist_y_rot

    if settings['show_figures']:
        ax.scatter(dist_x_rot, dist_y_rot, s=100, c='b', label='landing position')

    dist3 = max([abs(dist_x), abs(dist_y)])
    dist_rot3 = max([abs(dist_x_rot), abs(dist_y_rot)])

    factor = EDGE_SIZE_Y/EDGE_SIZE_X
    
    # print factor

    if abs(dist_y_rot) > factor * abs(dist_x_rot):
        dist_y = abs(dist_y_rot)
        dist_x = abs(dist_y_rot) / factor
        full_score_edge_size = FULL_SCORE_EDGE_SIZE_Y
        edge_size = EDGE_SIZE_Y
    else:
        dist_y = abs(dist_x_rot) * factor
        dist_x = abs(dist_x_rot)
        full_score_edge_size = FULL_SCORE_EDGE_SIZE_X
        edge_size = EDGE_SIZE_X
        
    eval_m = -PLATFORM_MAX_SCORE/(edge_size - full_score_edge_size)
    eval_q = PLATFORM_MAX_SCORE - full_score_edge_size * eval_m

    if settings['show_figures']:
        ax.plot([- dist_x, + dist_x], [- dist_y, - dist_y], 'b--')
        ax.plot([- dist_x, + dist_x], [+ dist_y, + dist_y], 'b--')
        ax.plot([+ dist_x, + dist_x], [- dist_y, + dist_y], 'b--')
        ax.plot([- dist_x, - dist_x], [- dist_y, + dist_y], 'b--')
        
        
    
        ax.legend(bbox_to_anchor=(0., 1.02, 1., .102), loc=3,
               ncol=1, mode="expand", borderaxespad=0.)
        ax.grid()

    if dist3 < full_score_edge_size:
        score_tr = 100
    elif dist3 > edge_size:
        score_tr = 0
    else:
        score_tr = (dist3*eval_m+eval_q)

    if 1:  # managed
        str1 = 'you landed on the truck after after {0:.2f} seconds'.format(t_land3)
        str2 = 'x distance from truck center is {0:.2f} meters'.format(dist_x_rot)
        str3 = 'y distance from truck center is {0:.2f} meters'.format(dist_y_rot)
        str4 = 'you scored {0:.2f} accuracy points'.format(score_tr)

        perf['accuracy_score_task3'] = score_tr

        if score_tr == 0:
            perf['time_task3'] = 600
        else:
            perf['time_task3'] = t_land3 - perf['time_task2'] - perf['time_task1']
    else:
        str1 = 'you did not manage to land on the plaform before timeout'
        str2 = 'you scored 0 accuracy points'
        str3 = ''
        str4 = ''

    print '\n'
    print '\n'
    print str1
    print str2
    print str3
    print str4
    print '\n'


    if settings['write_txt'] or settings['write_spreadsheet']:
        file_n = 'TP_3_eval_of_{}'.format(info['date_time'])


    if settings['write_txt']:
        file_name = open(file_n + '.txt', 'w')
        # file_name.write(str0)

        file_name.write('\n')
        file_name.write('\n')
        file_name.write(str1 + '\n')
        file_name.write(str2 + '\n')
        file_name.write(str3 + '\n')
        file_name.write('\n')

        file_name.close()

    if settings['write_spreadsheet']:
        # Create a workbook and add a worksheet.

        workbook = xlsxwriter.Workbook(file_n + '.xlsx')
        worksheet = workbook.add_worksheet()

        # Add a bold format to use to highlight cells.

        bold = workbook.add_format({'bold': True})

        # Write some data headers.

        worksheet.write('A1', 'time [s]', bold)
        worksheet.write('B1', 'x coordinate [m]', bold)
        worksheet.write('C1', 'y coordinate [m]', bold)
        worksheet.write('D1', 'z coordinate [m]', bold)

        worksheet.write('G1', 'x (truck) [m]', bold)
        worksheet.write('H1', 'y (truck) [m]', bold)
        worksheet.write('I1', 'z (truck) [m]', bold)

        worksheet.write('K1', 'vx (truck) [m]', bold)
        worksheet.write('L1', 'vy (truck) [m]', bold)
        worksheet.write('M1', 'vz (truck) [m]', bold)


        # Start from the first cell below the headers.

        row = 1
        col = 6

        # Iterate over the data and write it out row by row.

        for item in range(len(pxarr.tolist())):
            worksheet.write(row, col,     tr_xarr.tolist()[item] if not math.isnan(tr_xarr.tolist()[item]) else 0)
            worksheet.write(row, col + 1,     tr_yarr.tolist()[item] if not math.isnan(tr_yarr.tolist()[item]) else 0)
            worksheet.write(row, col + 2,     tr_zarr.tolist()[item] if not math.isnan(tr_zarr.tolist()[item]) else 0)
            row += 1

        row = 1
        col = 10

        # Iterate over the data and write it out row by row.

        for item in range(len(pxarr.tolist())):
            worksheet.write(row, col,     tr_xarr.tolist()[item] if not math.isnan(tr_xarr.tolist()[item]) else 0)
            worksheet.write(row, col + 1,     tr_yarr.tolist()[item] if not math.isnan(tr_yarr.tolist()[item]) else 0)
            worksheet.write(row, col + 2,     tr_zarr.tolist()[item] if not math.isnan(tr_zarr.tolist()[item]) else 0)
            row += 1

        # Start from the first cell below the headers.

        row = 1
        col = 0


        # Iterate over the data and write it out row by row.

        for item in range(len(pxarr.tolist())):
            worksheet.write(row, col,     tarr.tolist()[item] if not math.isnan(tarr.tolist()[item]) else 0)
            worksheet.write(row, col + 1,     pxarr.tolist()[item] if not math.isnan(pxarr.tolist()[item]) else 0)
            worksheet.write(row, col + 2,     pyarr.tolist()[item] if not math.isnan(pyarr.tolist()[item]) else 0)
            worksheet.write(row, col + 3,     pzarr.tolist()[item] if not math.isnan(pzarr.tolist()[item]) else 0)
            row += 1

        workbook.close()


def fill_evaluation_spreadsheet(perf):
    # Folder of this script
    source_folder = os.path.dirname(os.path.abspath(__file__))

    # Output file
    evaluation_filename = source_folder + '/evaluation.xlsx' 

    sciper = 'dummy SCIPER'  # dummy


    filename = source_folder + '/sciper.txt'

    f = open(filename, 'r')

    a = []

    for line in f:
        a.append(line[:-1])

    print 'sciper = ', a[0]

    sciper = a[0]

#    # Open output file
#    book = openpyxl.load_workbook(evaluation_filename)
#    sheet = book.active
#
#    # Add a bold format to use to highlight cells.
#
#    # Write some data headers.
#    idx = 1
#    empty = sheet.cell(row=1, column=idx).value == None
#
#    while not empty:
#        idx =idx+1
#        if sheet.cell(row=1, column=idx).value == None:
#            empty = 1
#
#    # Start from the first cell below the headers.
#    row = 1
#    col = idx
#    sheet.cell(column=col, row=row, value=sciper)
#
#    # Iterate over the data and write it out row by row.
#    sheet.cell(column=col, row=row + 1, value=perf['score_waypoints'][0]) if len(perf['score_waypoints']) >= 1 else 0
#    sheet.cell(column=col, row=row + 2, value=perf['score_waypoints'][1]) if len(perf['score_waypoints']) >= 2 else 0
#    sheet.cell(column=col, row=row + 3, value=perf['score_waypoints'][2]) if len(perf['score_waypoints']) >= 3 else 0
#    sheet.cell(column=col, row=row + 4, value=perf['accuracy_score_task1'])
#    sheet.cell(column=col, row=row + 5, value=perf['time_task1'])
#    sheet.cell(column=col, row=row + 6, value=perf['accuracy_score_task2'])
#    sheet.cell(column=col, row=row + 7, value=perf['time_task2'])
#    sheet.cell(column=col, row=row + 8, value=perf['accuracy_score_task3'])
#    sheet.cell(column=col, row=row + 9, value=perf['time_task3'])
#    sheet.cell(column=col, row=row + 12, value=perf['accuracy_score_task1'] + perf['accuracy_score_task2'] + perf['accuracy_score_task3'])
#    sheet.cell(column=col, row=row + 14, value=perf['time_task1'] + perf['time_task2'] + perf['time_task3'])

#    if perf['task1_not_done']:
#        sheet.cell(column=col, row=row + 1, value='x')
#        sheet.cell(column=col, row=row + 2, value='x')
#        sheet.cell(column=col, row=row + 3, value='x')
#        sheet.cell(column=col, row=row + 4, value='x')
#        sheet.cell(column=col, row=row + 5, value='x')
#
#    if perf['task2_not_done']:
#        sheet.cell(column=col, row=row + 6, value='x')
#        sheet.cell(column=col, row=row + 7, value='x')
#    if perf['task3_not_done']:
#        sheet.cell(column=col, row=row + 8, value='x')
#        sheet.cell(column=col, row=row + 9, value='x')
#
#    if perf['task1_not_done'] or perf['task2_not_done'] or perf['task3_not_done']:
#        sheet.cell(column=col, row=row + 12, value='x')
#        sheet.cell(column=col, row=row + 14, value='x')

    # timeout and 0 accuracy if not done

    if perf['task1_not_done']:
        perf['accuracy_score_task1'] = 0
        perf['time_task1'] = 600
    if perf['task2_not_done']:
        perf['accuracy_score_task2'] = 0
        perf['time_task2'] = 600
    if perf['task3_not_done']:
        perf['accuracy_score_task3'] = 0
        perf['time_task3'] = 600

    # timeout if 0 accuracy

    if perf['accuracy_score_task1'] == 0:
        perf['time_task1'] = 600
    if perf['accuracy_score_task2'] == 0:
        perf['time_task2'] = 600
    if perf['accuracy_score_task3'] == 0:
        perf['time_task3'] = 600

    filename = (source_folder + '/scores_' + sciper + '.txt');
    

    header =np.char.array(['sciper', 'score_waypoints_1', 'score_waypoints_2', 'score_waypoints_3', 'accuracy_score_task1', 'time_task1', 'accuracy_score_task2', 'time_task2', 'accuracy_score_task3', 'time_task3', 'accuracy_score_TOT', 'time_TOT'])
    scores =np.array([int(sciper), perf['score_waypoints'][0], perf['score_waypoints'][1], perf['score_waypoints'][2], perf['accuracy_score_task1'], perf['time_task1'], perf['accuracy_score_task2'], perf['time_task2'], perf['accuracy_score_task3'], perf['time_task3'], perf['accuracy_score_task1'] + perf['accuracy_score_task2'] + perf['accuracy_score_task3'], perf['time_task1'] + perf['time_task2'] + perf['time_task3']])

    data = np.vstack([header,scores])
    
    if os.path.isfile(filename):
        my_data = np.genfromtxt(filename, delimiter=',')
        data = np.vstack([my_data,scores])
            
    np.savetxt((source_folder + '/scores_' + sciper + '.txt'), (data), delimiter=",", fmt="%s")
    np.savetxt((source_folder + '/scores.txt'), (data), delimiter=",", fmt="%s")

    # Save and close file
#    book.save(evaluation_filename)
#    book.close()


def main():

    # p_opt = optparse.OptionParser()
    # p_opt.add_option('-t', '--task', dest='task_n', help='The task you want to evaluate', type=int)

    p = argparse.ArgumentParser()
    _ = p.add_argument('-s', '--spr', action='store_true', default=None, help='Write spreadsheet')
    _ = p.add_argument('-l', '--log', action='store_true', default=None, help='Write log tile')
    _ = p.add_argument('-e', '--eval', action='store_true', default=None, help='Fill evaluation spreadsheet')
    _ = p.add_argument('-f', '--fig', action='store_true', default=None, help='Display figures')
    _ = p.add_argument('-d', '--dev', action='store_true', help='Evaluate developer version ')
    _ = p.add_argument('-t', '--task', action='store', help='The task you want to evaluate')


    # (options, args1) = p_opt.parse_args()
    (args) = p.parse_args()


    settings = {'eval_task_1': True,
                'eval_task_2': True,
                'eval_task_3': True,
                'write_txt': False,
                'write_spreadsheet': False,
                'show_figures': False,
                'dev': False,
                'fill_eval_spreadsheet': False}


    if args.task == '1':
        settings['eval_task_2'] = False
        settings['eval_task_3'] = False
    elif args.task == '2':
        settings['eval_task_1'] = False
        settings['eval_task_3'] = False
    elif args.task == '3':
        settings['eval_task_1'] = False
        settings['eval_task_2'] = False

    if args.log:
        settings['write_txt'] = True
    if args.spr:
        settings['write_spreadsheet'] = True
    if args.eval:
        settings['fill_eval_spreadsheet'] = True
    if args.fig:
        settings['show_figures'] = True
    if args.dev:
        settings['dev'] = True


    # timestamp

    info = {'date_time': datetime.datetime.now()}

    data = get_data_from_log(settings)

    perf = {'cross_time_waypoints': [],
            'min_dist_waypoints': [],
            'score_waypoints': [],
            'accuracy_score_task1': 0,
            'time_task1': 0,
            'accuracy_score_task2': 0,
            'time_task2': 0,
            'accuracy_score_task3': 0,
            'time_task3': 0,
            'task1_not_done': 0,
            'task2_not_done': 0,
            'task3_not_done': 0}


    # # # # # # # # # #         Task1         # # # # # # # # # #


    if settings['eval_task_1']:
        evaluation_task_1(data, perf, settings, info)


    # # # # # # # # # #         Task2         # # # # # # # # # #


    if settings['eval_task_2']:
        evaluation_task_2(data, perf, settings, info)


    # # # # # # # # # #         Task 3         # # # # # # # # # #


    if settings['eval_task_3']:
        evaluation_task_3(data, perf, settings, info)


    # # # # # # # # # #         PERFORMANCE SPREADSHEET         # # # # # # # # # #


    if settings['fill_eval_spreadsheet']:
        fill_evaluation_spreadsheet(perf)


    # # # # # # # # # #              SHOW FIGURES               # # # # # # # # # #


    if settings['show_figures']:
        plt.show()


if __name__ == '__main__':
    main()

#
